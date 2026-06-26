#!/usr/bin/env python3
import base64
import hashlib
import html
import json
import os
import re
import sys
import tempfile
from datetime import date, datetime, time
from itertools import zip_longest
from pathlib import Path
from typing import Any


def main() -> int:
    try:
        request = json.load(sys.stdin)
        result = parse_with_openpyxl(request)
        print(json.dumps(result, ensure_ascii=False))
        return 0
    except ModuleNotFoundError as error:
        if error.name == "openpyxl":
            print(
                "openpyxl is not installed. Install locally with `python3 -m pip install openpyxl` "
                "or set RAG_OPENPYXL_PYTHON to a Python environment that has openpyxl.",
                file=sys.stderr,
            )
            return 2
        raise
    except Exception as error:
        print(str(error), file=sys.stderr)
        return 1


def parse_with_openpyxl(request: dict[str, Any]) -> dict[str, Any]:
    from openpyxl import load_workbook

    source_path, cleanup_dir = materialize_source(request)
    try:
        workbook = load_workbook(source_path, data_only=True, read_only=True)
        formula_workbook = load_workbook(source_path, data_only=False, read_only=True)
        asset_workbook = load_workbook(source_path, data_only=False, read_only=False)
        asset_dir = asset_output_directory(request, source_path)
        sheets = []
        body_parts = []
        regions = []
        tables = []
        visual_assets = []
        warnings = []
        offset = 0
        for sheet_index, (sheet, formula_sheet) in enumerate(
            zip(workbook.worksheets, formula_workbook.worksheets), start=1
        ):
            rows, row_warnings = materialized_rows(sheet, formula_sheet)
            warnings.extend(row_warnings)
            if not rows:
                continue
            sheet_state = str(getattr(formula_sheet, "sheet_state", "visible") or "visible")
            sheet_title = f"# {sheet.title}"
            body_parts.append(sheet_title)
            title_start = offset
            offset += len(sheet_title)
            regions.append(
                {
                    "id": f"sheet_{sheet_index}_title",
                    "kind": "heading",
                    "pageNumber": sheet_index,
                    "text": sheet_title,
                    "characterStart": title_start,
                    "characterEnd": offset,
                    "metadata": {"sheetName": sheet.title, "sheetState": sheet_state},
                }
            )
            offset += 2

            table_text = "\n".join(" | ".join(cell for cell in row["values"]) for row in rows)
            body_parts.append(table_text)
            table_start = offset
            offset += len(table_text)
            region_id = f"sheet_{sheet_index}_table_region"
            regions.append(
                {
                    "id": region_id,
                    "kind": "table",
                    "pageNumber": sheet_index,
                    "text": table_text,
                    "characterStart": table_start,
                    "characterEnd": offset,
                    "metadata": {"sheetName": sheet.title, "sheetState": sheet_state},
                }
            )
            offset += 2

            tables.append(
                {
                    "id": f"sheet_{sheet_index}_table",
                    "pageNumber": sheet_index,
                    "regionId": region_id,
                    "cells": table_cells(rows),
                    "summary": table_text[:500],
                    "metadata": {
                        "sheetName": sheet.title,
                        "sheetState": sheet_state,
                        "maxRow": sheet.max_row,
                        "maxColumn": sheet.max_column,
                    },
                }
            )
            sheets.append(
                {
                    "name": sheet.title,
                    "state": sheet_state,
                    "pageNumber": sheet_index,
                    "rowCount": len(rows),
                    "columnCount": max_width(rows),
                }
            )
            asset_sheet = asset_workbook[sheet.title] if sheet.title in asset_workbook.sheetnames else None
            if asset_sheet is not None:
                visual_assets.extend(
                    visual_assets_for_sheet(
                        asset_sheet=asset_sheet,
                        page_number=sheet_index,
                        sheet_name=sheet.title,
                        sheet_state=sheet_state,
                        asset_dir=asset_dir,
                    )
                )

        body = "\n\n".join(body_parts)
        layout = {
            "parserId": "openpyxl-rag-parser",
            "strategy": "table_structure",
            "pages": [
                {
                    "pageNumber": sheet["pageNumber"],
                    "width": max(1, sheet["columnCount"]),
                    "height": max(1, sheet["rowCount"]),
                    "unit": "normalized",
                    "metadata": {"sheetName": sheet["name"], "sheetState": sheet["state"]},
                }
                for sheet in sheets
            ]
            or [{"pageNumber": 1, "width": 1, "height": 1, "unit": "normalized"}],
            "regions": regions
            or [
                {
                    "id": "empty_workbook",
                    "kind": "table",
                    "pageNumber": 1,
                    "text": body,
                    "characterStart": 0,
                    "characterEnd": len(body),
                }
            ],
            "tables": tables,
            "visualAssets": visual_assets,
            "metadata": {
                "sourceId": str(request.get("sourceId") or ""),
                "normalizer": "openpyxl-rag-parser",
                "sheetCount": len(sheets),
                "visualAssetCount": len(visual_assets),
            },
        }
        return {
            "body": body,
            "layout": layout,
            "metadata": {
                "engine": "openpyxl",
                "sheetCount": len(sheets),
                "visualAssetCount": len(visual_assets),
            },
            "warnings": warnings
            if sheets
            else [{"code": "empty_workbook", "message": "No non-empty sheets were parsed."}],
        }
    finally:
        if cleanup_dir is not None:
            cleanup_dir.cleanup()


def materialize_source(request: dict[str, Any]) -> tuple[Path, tempfile.TemporaryDirectory[str] | None]:
    path = request.get("path")
    if isinstance(path, str) and path and Path(path).exists():
        return Path(path), None

    origin_uri = request.get("originUri")
    if isinstance(origin_uri, str) and origin_uri.startswith("file://"):
        file_path = Path(origin_uri.removeprefix("file://"))
        if file_path.exists():
            return file_path, None

    cleanup_dir = tempfile.TemporaryDirectory(prefix="openpyxl-rag-")
    temp_dir = Path(cleanup_dir.name)
    bytes_base64 = request.get("bytesBase64")
    if isinstance(bytes_base64, str) and bytes_base64:
        source_path = temp_dir / "source.xlsx"
        source_path.write_bytes(base64.b64decode(bytes_base64))
        return source_path, cleanup_dir

    raise ValueError("OpenPyXL parser requires path, file:// originUri, or bytesBase64.")


def materialized_rows(value_sheet: Any, formula_sheet: Any) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    rows = []
    warnings = []
    value_rows = value_sheet.iter_rows(values_only=True)
    formula_rows = formula_sheet.iter_rows(values_only=False)
    for row_number, (value_row, formula_row) in enumerate(zip_longest(value_rows, formula_rows), start=1):
        values = []
        value_row = value_row or []
        formula_row = formula_row or []
        for column_number, (value, formula_cell) in enumerate(
            zip_longest(value_row, formula_row), start=1
        ):
            text = cell_text(value)
            formula_text = cell_text(getattr(formula_cell, "value", None))
            if text == "" and formula_text.startswith("="):
                text = formula_text
                warnings.append(
                    {
                        "code": "formula_without_cached_value",
                        "message": (
                            f"Formula at {formula_sheet.title}!R{row_number}C{column_number} "
                            "had no cached calculated value; emitted formula text."
                        ),
                    }
                )
            values.append(text)
        while values and values[-1] == "":
            values.pop()
        if any(value != "" for value in values):
            rows.append({"rowNumber": row_number, "values": values})
    return rows, warnings


def table_cells(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cells = []
    for row_index, row in enumerate(rows):
        for column_index, text in enumerate(row["values"]):
            cells.append(
                {
                    "rowIndex": row_index,
                    "columnIndex": column_index,
                    "sourceRowNumber": row["rowNumber"],
                    "sourceColumnNumber": column_index + 1,
                    "text": text,
                }
            )
    return cells


def max_width(rows: list[dict[str, Any]]) -> int:
    return max((len(row["values"]) for row in rows), default=0)


def visual_assets_for_sheet(
    asset_sheet: Any,
    page_number: int,
    sheet_name: str,
    sheet_state: str,
    asset_dir: Path,
) -> list[dict[str, Any]]:
    assets = []
    for index, chart in enumerate(getattr(asset_sheet, "_charts", []) or [], start=1):
        asset_id = f"sheet_{page_number}_chart_{index}"
        metadata = {
            "sheetName": sheet_name,
            "sheetState": sheet_state,
            "assetType": "chart",
            "chartType": type(chart).__name__,
        }
        title = chart_title(chart)
        if title:
            metadata["title"] = title
        anchor_cell = anchor_to_cell(getattr(chart, "anchor", None))
        if anchor_cell:
            metadata["anchorCell"] = anchor_cell
        svg = chart_svg(metadata)
        artifact_path = write_asset(asset_dir, f"{asset_id}.svg", svg)
        assets.append(
            {
                "id": asset_id,
                "kind": "figure",
                "pageNumber": page_number,
                "mediaType": "image/svg+xml",
                "uri": artifact_path.as_uri(),
                "checksum": hashlib.sha256(svg).hexdigest(),
                "metadata": {
                    **metadata,
                    "artifactKind": "generated_chart_svg",
                    "sourceMediaType": "application/vnd.openxmlformats-officedocument.drawingml.chart+xml",
                },
            }
        )

    for index, image in enumerate(getattr(asset_sheet, "_images", []) or [], start=1):
        asset_id = f"sheet_{page_number}_image_{index}"
        metadata = {
            "sheetName": sheet_name,
            "sheetState": sheet_state,
            "assetType": "image",
            "imagePath": str(getattr(image, "path", "") or ""),
        }
        anchor_cell = anchor_to_cell(getattr(image, "anchor", None))
        if anchor_cell:
            metadata["anchorCell"] = anchor_cell
        image_bytes = image_data(image)
        media_type = image_media_type(image)
        extension = extension_for_media_type(media_type)
        artifact_path = write_asset(asset_dir, f"{asset_id}{extension}", image_bytes)
        assets.append(
            {
                "id": asset_id,
                "kind": "figure",
                "pageNumber": page_number,
                "mediaType": media_type,
                "uri": artifact_path.as_uri(),
                "checksum": hashlib.sha256(image_bytes).hexdigest(),
                "metadata": {**metadata, "artifactKind": "embedded_image"},
            }
        )

    return assets


def asset_output_directory(request: dict[str, Any], source_path: Path) -> Path:
    base_dir = Path(os.environ.get("RAG_OPENPYXL_ASSET_DIR", ".rag/parser-assets/openpyxl"))
    relative_path = ""
    metadata = request.get("metadata")
    if isinstance(metadata, dict):
        relative_path = str(metadata.get("relativePath") or "")
    source_id = sanitize_path_part(str(request.get("sourceId") or "source"))
    stem = sanitize_path_part(Path(relative_path).stem or source_path.stem or "workbook")
    digest = hashlib.sha256(source_path.read_bytes()).hexdigest()[:16]
    directory = (base_dir / source_id / f"{stem}-{digest}").resolve()
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def sanitize_path_part(value: str) -> str:
    sanitized = re.sub(r"[^A-Za-z0-9_.-]+", "-", value).strip("-._")
    return sanitized[:80] or "asset"


def write_asset(asset_dir: Path, filename: str, data: bytes) -> Path:
    target = (asset_dir / filename).resolve()
    if not str(target).startswith(str(asset_dir.resolve())):
        raise ValueError("Asset path escaped the configured asset directory.")
    target.write_bytes(data)
    return target


def chart_svg(metadata: dict[str, Any]) -> bytes:
    title = html.escape(str(metadata.get("title") or "Spreadsheet chart"))
    chart_type = html.escape(str(metadata.get("chartType") or "chart"))
    sheet_name = html.escape(str(metadata.get("sheetName") or "sheet"))
    anchor_cell = html.escape(str(metadata.get("anchorCell") or "unknown anchor"))
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="640" height="360" viewBox="0 0 640 360" role="img" aria-label="{title}">
  <rect width="640" height="360" fill="#ffffff"/>
  <rect x="24" y="24" width="592" height="312" fill="#f8fafc" stroke="#334155" stroke-width="2"/>
  <text x="48" y="80" font-family="Arial, sans-serif" font-size="28" font-weight="700" fill="#0f172a">{title}</text>
  <text x="48" y="124" font-family="Arial, sans-serif" font-size="18" fill="#334155">Type: {chart_type}</text>
  <text x="48" y="154" font-family="Arial, sans-serif" font-size="18" fill="#334155">Sheet: {sheet_name}</text>
  <text x="48" y="184" font-family="Arial, sans-serif" font-size="18" fill="#334155">Anchor: {anchor_cell}</text>
  <rect x="60" y="228" width="90" height="68" fill="#2563eb"/>
  <rect x="176" y="198" width="90" height="98" fill="#16a34a"/>
  <rect x="292" y="168" width="90" height="128" fill="#f97316"/>
  <line x1="48" y1="296" x2="432" y2="296" stroke="#475569" stroke-width="2"/>
</svg>
"""
    return svg.encode("utf-8")


def chart_title(chart: Any) -> str:
    title = getattr(chart, "title", None)
    if title is None:
        return ""
    try:
        paragraphs = title.tx.rich.p
        parts = []
        for paragraph in paragraphs:
            for run in getattr(paragraph, "r", []) or []:
                value = getattr(run, "t", "")
                if value:
                    parts.append(str(value))
        return " ".join(parts).strip()
    except Exception:
        return ""


def anchor_to_cell(anchor: Any) -> str:
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return ""
    row = getattr(marker, "row", None)
    col = getattr(marker, "col", None)
    if row is None or col is None:
        return ""
    return f"R{int(row) + 1}C{int(col) + 1}"


def image_media_type(image: Any) -> str:
    image_format = str(getattr(image, "format", "") or "").lower()
    if image_format in {"jpeg", "jpg"}:
        return "image/jpeg"
    if image_format == "gif":
        return "image/gif"
    if image_format == "bmp":
        return "image/bmp"
    return "image/png"


def extension_for_media_type(media_type: str) -> str:
    if media_type == "image/jpeg":
        return ".jpg"
    if media_type == "image/gif":
        return ".gif"
    if media_type == "image/bmp":
        return ".bmp"
    return ".png"


def image_data(image: Any) -> bytes:
    try:
        data = image._data()
        if isinstance(data, bytes):
            return data
    except Exception:
        pass
    fallback = json.dumps(
        {"imagePath": str(getattr(image, "path", "") or ""), "format": str(getattr(image, "format", "") or "")},
        sort_keys=True,
    )
    return fallback.encode("utf-8")


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)


if __name__ == "__main__":
    raise SystemExit(main())
